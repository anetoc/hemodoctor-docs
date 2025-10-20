#!/usr/bin/env python3
"""
Gera documento técnico para dev team com diagramas Mermaid
"""
import openpyxl
from datetime import datetime

def generate_technical_document():
    wb = openpyxl.load_workbook('HEMODOCTOR_REGRAS_COMPLETAS_v2.4.0.xlsx', data_only=True)

    md = []
    md.append("# HemoDoctor Hybrid v2.4.0")
    md.append("## Especificacao Tecnica para Dev Team\n")
    md.append(f"**Data:** {datetime.now().strftime('%Y-%m-%d')}")
    md.append("**Versao:** v2.4.0")
    md.append("**Status:** Ready for Sprint 0 Implementation\n")
    md.append("---\n")

    # Executive Summary
    md.append("## EXECUTIVE SUMMARY\n")
    md.append("Sistema de apoio a decisao clinica baseado em regras deterministicas YAML.")
    md.append("**Stack:** FastAPI + Python 3.11+ + YAML + pytest")
    md.append("**Deployment:** Docker + Kubernetes")
    md.append("**Compliance:** ANVISA RDC 657/2022, FDA 21 CFR Part 11, ISO 13485\n")

    # Architecture
    md.append("## ARQUITETURA DO SISTEMA\n")
    md.append("```mermaid")
    md.append("graph TD")
    md.append("    A[CBC Input API] --> B[Input Validator]")
    md.append("    B --> C[Normalization Engine]")
    md.append("    C --> D[Schema Validator]")
    md.append("    D --> E[Evidence Engine 79 rules]")
    md.append("    E --> F[Syndrome DAG Fusion 35 syndromes]")
    md.append("    F --> G{Criticality}")
    md.append("    G -->|Critical| H[Short-Circuit Handler]")
    md.append("    G -->|Priority| I[Full Analysis]")
    md.append("    G -->|Routine| J[Standard Output]")
    md.append("    H --> K[Next Steps Engine]")
    md.append("    I --> K")
    md.append("    J --> K")
    md.append("    K --> L[Output Renderer]")
    md.append("    L --> M[WORM Logger]")
    md.append("    M --> N[API Response]")
    md.append("```\n")

    # Modules
    md.append("## MODULOS PRINCIPAIS\n")
    md.append("### 1. Input Validator")
    md.append("- **File:** `src/api/input_validator.py`")
    md.append("- **Schema:** `01_schema_hybrid.yaml` (42 campos)")
    md.append("- **Validacoes:** Type checking, range validation, required fields")
    md.append("- **Output:** Canonical CBC dict\n")

    md.append("### 2. Evidence Engine")
    md.append("- **File:** `src/engines/evidence_engine.py`")
    md.append("- **Config:** `02_evidence_hybrid.yaml` (79 evidencias)")
    md.append("- **Metodo:** simpleeval (NUNCA usar eval() direto)")
    md.append("- **Output:** List[Evidence] com status (present/absent/unknown)\n")

    md.append("### 3. Syndrome DAG Fusion")
    md.append("- **File:** `src/engines/syndrome_engine.py`")
    md.append("- **Config:** `03_syndromes_hybrid.yaml` (35 sindromes)")
    md.append("- **Logica:** ALL/ANY/NEGATIVE + threshold")
    md.append("- **Short-circuit:** Critical syndromes stop further processing")
    md.append("- **Output:** List[Syndrome] ordenado por criticidade\n")

    md.append("### 4. Next Steps Engine")
    md.append("- **File:** `src/engines/next_steps_engine.py`")
    md.append("- **Config:** `09_next_steps_engine_hybrid.yaml` (40 triggers)")
    md.append("- **Logica:** When conditions (Python expressions)")
    md.append("- **Output:** Prioritized list of clinical recommendations\n")

    md.append("### 5. WORM Logger")
    md.append("- **File:** `src/audit/worm_logger.py`")
    md.append("- **Config:** `08_wormlog_hybrid.yaml`")
    md.append("- **Features:** HMAC-SHA256, immutable append-only log")
    md.append("- **Retention:** 1825 dias (5 anos)")
    md.append("- **Compliance:** FDA 21 CFR Part 11, LGPD Art. 16\n")

    # Data Flow
    md.append("## DATA FLOW DETALHADO\n")
    md.append("```mermaid")
    md.append("sequenceDiagram")
    md.append("    participant Client")
    md.append("    participant API")
    md.append("    participant Validator")
    md.append("    participant EvidenceEngine")
    md.append("    participant SyndromeEngine")
    md.append("    participant NextStepsEngine")
    md.append("    participant WORMLogger")
    md.append("    Client->>API: POST /analyze (CBC JSON)")
    md.append("    API->>Validator: validate_input(cbc)")
    md.append("    Validator->>API: canonical_cbc")
    md.append("    API->>EvidenceEngine: evaluate_evidences(cbc)")
    md.append("    EvidenceEngine->>API: List[Evidence] (79 max)")
    md.append("    API->>SyndromeEngine: fuse_syndromes(evidences)")
    md.append("    SyndromeEngine->>API: List[Syndrome] (1-5 typical)")
    md.append("    API->>NextStepsEngine: recommend_next_steps(syndromes)")
    md.append("    NextStepsEngine->>API: List[NextStep]")
    md.append("    API->>WORMLogger: log_analysis(cbc, evidences, syndromes)")
    md.append("    WORMLogger->>API: log_id + hmac")
    md.append("    API->>Client: JSON Response + route_id")
    md.append("```\n")

    # Schema
    md.append("## SCHEMA CBC (42 campos)\n")
    schema_ws = wb['Schema CBC']
    md.append("| Campo | Tipo | Unidade | Required | Range |")
    md.append("|-------|------|---------|----------|-------|")
    for row in schema_ws.iter_rows(min_row=2, max_row=15, values_only=True):
        if row[0]:
            md.append(f"| {row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} |")
    md.append("\n*Ver Excel para todos os 42 campos*\n")

    # Evidences (amostra)
    md.append("## EVIDENCIAS (79 regras)\n")
    md.append("### Exemplo de Implementacao\n")
    md.append("```python")
    md.append("from simpleeval import simple_eval\n")
    md.append("def evaluate_evidence(evidence_rule: str, cbc: dict, config: dict) -> str:")
    md.append('    """Evaluate a single evidence rule')
    md.append('    Returns: "present" | "absent" | "unknown"')
    md.append('    """')
    md.append("    try:")
    md.append("        context = {**cbc, 'config': config}")
    md.append("        result = simple_eval(evidence_rule, names=context)")
    md.append('        return "present" if result else "absent"')
    md.append("    except (KeyError, AttributeError):")
    md.append('        return "unknown"  # Missing data')
    md.append("```\n")

    md.append("### Categorias de Evidencias\n")
    ev_ws = wb['Evidências']
    categorias = {}
    for row in ev_ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            cat = row[1]
            categorias[cat] = categorias.get(cat, 0) + 1

    for cat, count in sorted(categorias.items()):
        md.append(f"- **{cat}:** {count} evidencias")
    md.append("\n")

    # Syndromes (amostra)
    md.append("## SINDROMES (35 total)\n")
    md.append("### Distribuicao por Criticidade\n")
    syn_ws = wb['Síndromes']
    criticidades = {}
    for row in syn_ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            crit = row[1]
            criticidades[crit] = criticidades.get(crit, 0) + 1

    for crit, count in sorted(criticidades.items()):
        md.append(f"- **{crit}:** {count} sindromes")
    md.append("\n")

    md.append("### Exemplo de Logica de Fusao\n")
    md.append("```python")
    md.append("def fuse_syndrome(syndrome_def: dict, evidences: List[Evidence]) -> Optional[Syndrome]:")
    md.append('    """DAG fusion logic for syndromes"""')
    md.append("    evidence_ids = {e.id for e in evidences if e.status == 'present'}")
    md.append("    ")
    md.append("    # Check ALL requirements")
    md.append("    if syndrome_def['combine'].get('all'):")
    md.append("        if not all(req in evidence_ids for req in syndrome_def['combine']['all']):")
    md.append("            return None")
    md.append("    ")
    md.append("    # Check ANY requirements")
    md.append("    if syndrome_def['combine'].get('any'):")
    md.append("        if not any(req in evidence_ids for req in syndrome_def['combine']['any']):")
    md.append("            return None")
    md.append("    ")
    md.append("    # Check NEGATIVE exclusions")
    md.append("    if syndrome_def.get('negative'):")
    md.append("        if any(neg in evidence_ids for neg in syndrome_def['negative']):")
    md.append("            return None")
    md.append("    ")
    md.append("    return Syndrome(id=syndrome_def['id'], criticality=syndrome_def['criticality'])")
    md.append("```\n")

    # Testing
    md.append("## TESTES (Sprint 0 Target)\n")
    md.append("### Coverage Target: 85%\n")
    md.append("### Test Structure\n")
    md.append("```")
    md.append("tests/")
    md.append("├── unit/")
    md.append("│   ├── test_evidence_engine.py  (79 test cases)")
    md.append("│   ├── test_syndrome_engine.py  (35 test cases)")
    md.append("│   └── test_next_steps_engine.py (40 test cases)")
    md.append("├── integration/")
    md.append("│   ├── test_end_to_end.py")
    md.append("│   └── test_api_endpoints.py")
    md.append("└── fixtures/")
    md.append("    └── cbc_samples.yaml  (240+ casos Red List)")
    md.append("```\n")

    md.append("### Exemplo de Test Case\n")
    md.append("```python")
    md.append("def test_evidence_E_ANC_CRIT():")
    md.append('    """Test critical neutropenia detection"""')
    md.append("    cbc = {'anc': 0.3, 'age_months': 120, 'sex': 'M'}")
    md.append("    config = load_config('00_config_hybrid.yaml')")
    md.append("    ")
    md.append("    evidences = evaluate_evidences(cbc, config)")
    md.append("    ")
    md.append("    assert 'E-ANC-CRIT' in [e.id for e in evidences if e.status == 'present']")
    md.append("    assert evidences[0].strength == 'strong'")
    md.append("```\n")

    # API Spec
    md.append("## API SPECIFICATION\n")
    md.append("### POST /api/v1/analyze\n")
    md.append("**Request:**")
    md.append("```json")
    md.append("{")
    md.append('  "patient_id": "hash_or_pseudonym",')
    md.append('  "cbc": {')
    md.append('    "wbc": 4.5,')
    md.append('    "hb": 12.5,')
    md.append('    "plt": 200,')
    md.append('    "anc": 2.5,')
    md.append('    "age_months": 120,')
    md.append('    "sex": "M"')
    md.append("  }")
    md.append("}")
    md.append("```\n")

    md.append("**Response:**")
    md.append("```json")
    md.append("{")
    md.append('  "route_id": "sha256_hash_of_decision_path",')
    md.append('  "syndromes": [')
    md.append('    {')
    md.append('      "id": "S-NEUTROPENIA-GRAVE",')
    md.append('      "criticality": "critical",')
    md.append('      "confidence": 0.95')
    md.append("    }")
    md.append("  ],")
    md.append('  "next_steps": [')
    md.append('    {')
    md.append('      "level": "critical",')
    md.append('      "test": "Blood smear + Bone marrow biopsy",')
    md.append('      "turnaround": "<2h"')
    md.append("    }")
    md.append("  ],")
    md.append('  "worm_log_id": "log_entry_uuid"')
    md.append("}")
    md.append("```\n")

    # Implementation Roadmap
    md.append("## IMPLEMENTATION ROADMAP (Sprint 0)\n")
    md.append("### Week 1 (20-26 Out)\n")
    md.append("**Day 1-2:** Setup + Core Infrastructure")
    md.append("- [ ] FastAPI project structure")
    md.append("- [ ] YAML loaders (00-12 configs)")
    md.append("- [ ] Canonical schema validator (01_schema)")
    md.append("- [ ] pytest setup + fixtures\n")

    md.append("**Day 3-4:** Evidence + Syndrome Engines")
    md.append("- [ ] Evidence engine (79 regras)")
    md.append("- [ ] Syndrome DAG fusion (35 sindromes)")
    md.append("- [ ] Short-circuit logic (critical)")
    md.append("- [ ] Unit tests (85% coverage target)\n")

    md.append("**Day 5-6:** Next Steps + WORM")
    md.append("- [ ] Next steps engine (40 triggers)")
    md.append("- [ ] WORM logger + HMAC")
    md.append("- [ ] Integration tests")
    md.append("- [ ] API endpoints\n")

    md.append("**Day 7:** Testing + Documentation")
    md.append("- [ ] Red List validation prep (240 casos)")
    md.append("- [ ] API documentation (Swagger)")
    md.append("- [ ] Deployment configs (Docker)")
    md.append("- [ ] Sprint 0 review\n")

    # Security
    md.append("## SECURITY REQUIREMENTS\n")
    md.append("### Input Validation")
    md.append("- **NEVER use eval() directly** - always use simpleeval")
    md.append("- Validate all numeric ranges (prevent overflow)")
    md.append("- Sanitize all string inputs")
    md.append("- Rate limiting: 1000 req/hour per IP\n")

    md.append("### Data Protection")
    md.append("- Pseudonymization: SHA256(patient_id)")
    md.append("- Zero PHI in logs (only route_id + hashes)")
    md.append("- WORM log retention: 1825 days (5 years)")
    md.append("- Encryption at rest: AES-256-GCM\n")

    # Performance
    md.append("## PERFORMANCE TARGETS\n")
    md.append("- **Latency:** <100ms per analysis (p95)")
    md.append("- **Throughput:** >1000 cases/hour")
    md.append("- **Concurrency:** 50 simultaneous requests")
    md.append("- **Availability:** 99.9% uptime\n")

    # References
    md.append("## REFERENCIAS\n")
    md.append("- Excel completo: `HEMODOCTOR_REGRAS_COMPLETAS_v2.4.0.xlsx`")
    md.append("- YAMLs fonte: `HEMODOCTOR_HIBRIDO_V1.0/YAMLs/`")
    md.append("- QUICKSTART: `QUICKSTART_IMPLEMENTACAO.md`")
    md.append("- Runbook: `10_runbook_hybrid.yaml`")
    md.append("\n---\n")

    md.append(f"\n**Gerado em:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    md.append("**Versao:** v2.4.0")
    md.append("**Status:** Ready for Sprint 0 Implementation")

    # Salvar
    content = '\n'.join(md)
    with open('ESPECIFICACAO_TECNICA_DEV_TEAM_v2.4.0.md', 'w', encoding='utf-8') as f:
        f.write(content)

    print("✅ Documento tecnico gerado: ESPECIFICACAO_TECNICA_DEV_TEAM_v2.4.0.md")
    print(f"📄 Linhas: {len(md)}")
    print(f"📊 Tamanho: {len(content) / 1024:.1f} KB")

    return 'ESPECIFICACAO_TECNICA_DEV_TEAM_v2.4.0.md'

if __name__ == '__main__':
    generate_technical_document()
