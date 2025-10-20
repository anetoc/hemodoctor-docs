#!/usr/bin/env python3
"""
Gera documento de validação clínica para hematologista
"""
import openpyxl
from datetime import datetime

def generate_clinical_document():
    # Ler Excel
    wb = openpyxl.load_workbook('HEMODOCTOR_REGRAS_COMPLETAS_v2.4.0.xlsx', data_only=True)

    # Iniciar documento
    md = []
    md.append("# HemoDoctor Hybrid v2.4.0")
    md.append("## Documento de Validacao Clinica para Hematologista\n")
    md.append(f"**Data:** {datetime.now().strftime('%d/%m/%Y')}")
    md.append("**Versao:** v2.4.0\n")
    md.append("---\n")

    # Sumário
    md.append("## SUMARIO EXECUTIVO\n")
    md.append("Sistema de apoio a decisao clinica (SaMD Class III) para analise de hemogramas.")
    md.append("**35 sindromes hematologicas** usando **79 evidencias clinicas**.\n")
    md.append("---\n")

    # Componentes
    md.append("## COMPONENTES DO SISTEMA\n")
    resumo_ws = wb['Resumo']
    md.append("| Componente | Quantidade |")
    md.append("|------------|------------|")
    for row in resumo_ws.iter_rows(min_row=2, max_row=10, values_only=True):
        if row[0]:
            md.append(f"| {row[0]} | {row[1]} |")
    md.append("\n---\n")

    # Evidências
    md.append("## EVIDENCIAS CLINICAS\n")
    ev_ws = wb['Evidências']

    count = 0
    for row in ev_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        count += 1

        md.append(f"\n### {count}. {row[0]}")
        md.append(f"\n**Categoria:** {row[1]}")
        md.append(f"\n**Regra:** `{row[2]}`")
        md.append(f"\n**Forca:** {row[3]}")
        md.append(f"\n**Descricao:** {row[4]}")
        md.append(f"\n**Significancia Clinica:** {row[5]}")
        if len(row) > 7 and row[7]:
            md.append(f"\n**Fonte:** {row[7]}")

        md.append("\n**Validacao:**")
        md.append("- [ ] Criterio clinicamente adequado?")
        md.append("- [ ] Cutoff correto para populacao alvo?")
        md.append("- [ ] Faixas pediatricas necessarias?")
        md.append("\n**Comentarios do Hematologista:**")
        md.append("```")
        md.append("_________________________________________________________________")
        md.append("```\n")
        md.append("---\n")

        if count >= 79:  # Limitar para nao exceder
            break

    # Síndromes
    md.append("\n## SINDROMES HEMATOLOGICAS\n")
    syn_ws = wb['Síndromes']

    count = 0
    for row in syn_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        count += 1

        md.append(f"\n### {count}. {row[0]}")
        md.append(f"\n**Criticidade:** {row[1]}")
        md.append(f"\n**Logica:** {row[2]}")
        md.append(f"\n**Evidencias Requeridas:** {row[3]}")
        if row[4]:
            md.append(f"\n**Evidencias Opcionais:** {row[4]}")
        if row[5]:
            md.append(f"\n**Exclusoes:** {row[5]}")
        md.append(f"\n**Threshold:** {row[6]}")

        md.append("\n**Validacao:**")
        md.append("- [ ] Logica ALL/ANY correta?")
        md.append("- [ ] Evidencias suficientes?")
        md.append("- [ ] Threshold adequado?")
        md.append("\n**Comentarios do Hematologista:**")
        md.append("```")
        md.append("_________________________________________________________________")
        md.append("```\n")
        md.append("---\n")

        if count >= 35:
            break

    # Next Steps (amostra)
    md.append("\n## PROXIMOS PASSOS CLINICOS (Amostra)\n")
    ns_ws = wb['Next Steps']

    count = 0
    for row in ns_ws.iter_rows(min_row=2, max_row=12, values_only=True):
        if not row[0]:
            continue
        count += 1

        md.append(f"\n### {count}. {row[0]}")
        md.append(f"\n**Sindrome Alvo:** {row[1]}")
        md.append(f"\n**Nivel:** {row[3]}")
        md.append(f"\n**Recomendacoes:** {row[4]}")
        md.append(f"\n**Racional:** {row[5]}")

        md.append("\n**Validacao:**")
        md.append("- [ ] Recomendacoes apropriadas?")
        md.append("- [ ] Urgencia correta?")
        md.append("\n**Comentarios:**")
        md.append("```")
        md.append("_________________________________________________________________")
        md.append("```\n")
        md.append("---\n")

    md.append("\n*Ver Excel para todos os 40 triggers*\n")

    # Formulário final
    md.append("\n## FORMULARIO DE VALIDACAO FINAL\n")
    md.append("\n### Aprovacao Geral\n")
    md.append("- [ ] Todas as 79 evidencias foram revisadas")
    md.append("- [ ] Todas as 35 sindromes foram revisadas")
    md.append("- [ ] Todos os next steps foram revisados")
    md.append("\n### Decisao Final\n")
    md.append("- [ ] APROVADO sem restricoes")
    md.append("- [ ] APROVADO com ressalvas (detalhar acima)")
    md.append("- [ ] NAO APROVADO (detalhar motivos acima)")
    md.append("\n**Nome do Hematologista:** ___________________________")
    md.append("\n**CRM:** ___________________________")
    md.append("\n**Data:** ___________________________")
    md.append("\n**Assinatura:** ___________________________\n")

    # Salvar
    content = '\n'.join(md)
    with open('VALIDACAO_CLINICA_HEMATOLOGISTA_v2.4.0.md', 'w', encoding='utf-8') as f:
        f.write(content)

    print("✅ Documento clinico gerado: VALIDACAO_CLINICA_HEMATOLOGISTA_v2.4.0.md")
    print(f"📄 Linhas: {len(md)}")
    print(f"📊 Tamanho estimado: {len(content) / 1024:.1f} KB")

    return 'VALIDACAO_CLINICA_HEMATOLOGISTA_v2.4.0.md'

if __name__ == '__main__':
    generate_clinical_document()
