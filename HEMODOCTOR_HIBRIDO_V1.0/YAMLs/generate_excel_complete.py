#!/usr/bin/env python3
"""
Generate complete Excel workbook from HemoDoctor YAMLs
Version: 2.4.0
Date: 2025-10-19
"""

import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

# Configurações
YAML_DIR = Path("/Users/abelcosta/Documents/HemoDoctor/docs/HEMODOCTOR_HIBRIDO_V1.0/YAMLs")
OUTPUT_FILE = Path("/Users/abelcosta/Documents/HemoDoctor/docs/HEMODOCTOR_REGRAS_COMPLETAS_v2.4.0.xlsx")

# Estilos
HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
HEADER_FONT = Font(bold=True)
MONOSPACE_FONT = Font(name='Courier New', size=9)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def load_yaml(filename):
    """Carrega arquivo YAML"""
    filepath = YAML_DIR / filename
    with open(filepath, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)

def apply_header_style(ws, row=1):
    """Aplica estilo ao cabeçalho"""
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def auto_adjust_columns(ws, min_width=10, max_width=80):
    """Ajusta largura das colunas"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_summary_sheet(wb, stats):
    """Cria aba Resumo"""
    ws = wb.create_sheet("Resumo", 0)

    data = [
        ["HemoDoctor Hybrid v2.4.0 - Especificação Completa"],
        [""],
        ["Informação", "Valor"],
        ["Versão", "v2.4.0"],
        ["Data", "19 Out 2025"],
        ["Total Evidências", stats['total_evidences']],
        ["Total Síndromes", stats['total_syndromes']],
        ["Total Triggers", stats['total_triggers']],
        ["Total Campos CBC", stats['total_fields']],
        [""],
        ["Distribuição Evidências"],
        ["Critical", stats['critical_evidences']],
        ["Strong", stats['strong_evidences']],
        ["Moderate", stats['moderate_evidences']],
        ["Weak", stats['weak_evidences']],
        [""],
        ["Distribuição Síndromes"],
        ["Critical", stats['critical_syndromes']],
        ["Priority", stats['priority_syndromes']],
        ["Review Sample", stats['review_syndromes']],
        ["Routine", stats['routine_syndromes']],
    ]

    for row in data:
        ws.append(row)

    # Estilo título
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:B1')

    # Estilo headers
    for row in [3, 11, 17]:
        ws[f'A{row}'].font = HEADER_FONT
        ws[f'B{row}'].font = HEADER_FONT
        ws[f'A{row}'].fill = HEADER_FILL
        ws[f'B{row}'].fill = HEADER_FILL

    auto_adjust_columns(ws)

def extract_all_evidences(data):
    """Extrai TODAS as evidências de todas as categorias"""
    evidences = []

    # Categorias de evidências
    categories = [
        'critical_evidences',
        'red_blood_cell_evidences',
        'white_blood_cell_evidences',
        'platelet_evidences',
        'coagulation_evidences',
        'molecular_evidences',
        'supplementary_lab_evidences',
        'pre_analytical_evidences',
        'complementary_evidences'
    ]

    for category in categories:
        if category in data:
            for evidence in data[category]:
                evidence['category'] = category.replace('_evidences', '').replace('_', ' ').title()
                evidences.append(evidence)

    return evidences

def create_evidences_sheet(wb, data):
    """Cria aba Evidências"""
    ws = wb.create_sheet("Evidências")

    # Headers
    headers = ["ID", "Categoria", "Rule", "Strength", "Description",
               "Clinical Significance", "Requires", "Source", "Differential", "Next Steps"]
    ws.append(headers)
    apply_header_style(ws)

    # Extrair todas as evidências
    evidences = extract_all_evidences(data)

    # Dados
    for ev in evidences:
        requires = ", ".join(ev.get('requires', [])) if ev.get('requires') else ""
        differential = "; ".join(ev.get('differential', [])) if ev.get('differential') else ""
        next_steps = "; ".join(ev.get('next_steps', [])) if ev.get('next_steps') else ""

        row = [
            ev.get('id', ''),
            ev.get('category', ''),
            ev.get('rule', ''),
            ev.get('strength', ''),
            ev.get('description', ''),
            ev.get('clinical_significance', ''),
            requires,
            ev.get('source', ''),
            differential,
            next_steps
        ]
        ws.append(row)

        # Aplica fonte monospace para regras
        rule_cell = ws.cell(row=ws.max_row, column=3)
        rule_cell.font = MONOSPACE_FONT

    auto_adjust_columns(ws)
    return len(evidences)

def extract_all_syndromes(data):
    """Extrai TODAS as síndromes de todas as criticidades"""
    syndromes = []

    categories = [
        'critical_syndromes',
        'priority_syndromes',
        'review_sample_syndromes',
        'routine_syndromes'
    ]

    for category in categories:
        if category in data:
            criticality = category.replace('_syndromes', '')
            for syndrome in data[category]:
                syndrome['criticality_category'] = criticality
                syndromes.append(syndrome)

    return syndromes

def create_syndromes_sheet(wb, data):
    """Cria aba Síndromes"""
    ws = wb.create_sheet("Síndromes")

    # Headers
    headers = ["ID", "Criticality", "Combine Logic", "Required Evidences (ALL)",
               "Optional Evidences (ANY)", "Negative Evidences", "Threshold",
               "Description", "Actions", "Next Steps", "Short Circuit"]
    ws.append(headers)
    apply_header_style(ws)

    # Extrair todas as síndromes
    syndromes = extract_all_syndromes(data)

    # Dados
    for syn in syndromes:
        combine = syn.get('combine', {})

        # Processar combine (pode ter estruturas complexas)
        required = []
        optional = []

        if isinstance(combine.get('all'), list):
            required = [str(item) for item in combine['all']]

        if isinstance(combine.get('any'), list):
            for item in combine['any']:
                if isinstance(item, dict):
                    optional.append(str(item))
                else:
                    optional.append(str(item))

        negative = combine.get('negative', [])

        actions = "; ".join(syn.get('actions', []))
        next_steps = "; ".join(syn.get('next_steps', []))

        row = [
            syn.get('id', ''),
            syn.get('criticality', ''),
            str(combine),
            ", ".join(required),
            ", ".join(optional),
            ", ".join(negative) if negative else "",
            syn.get('threshold', ''),
            syn.get('description', ''),
            actions,
            next_steps,
            "YES" if syn.get('short_circuit') else "NO"
        ]
        ws.append(row)

    auto_adjust_columns(ws)
    return len(syndromes)

def create_next_steps_sheet(wb, data):
    """Cria aba Next Steps"""
    ws = wb.create_sheet("Next Steps")

    # Headers
    headers = ["Trigger ID", "Syndrome Target", "When Condition", "Priority Level",
               "Test Recommendations", "Rationale", "Cost Band", "Turnaround"]
    ws.append(headers)
    apply_header_style(ws)

    triggers = data.get('triggers', [])

    # Dados
    for trigger in triggers:
        suggest = trigger.get('suggest', [{}])[0]

        tests = []
        rationale = []
        for item in trigger.get('suggest', []):
            if item.get('test'):
                tests.append(item['test'])
            if item.get('rationale'):
                rationale.append(item['rationale'])

        row = [
            trigger.get('id', ''),
            trigger.get('syndrome', ''),
            trigger.get('when', ''),
            suggest.get('level', ''),
            "; ".join(tests),
            "; ".join(rationale),
            suggest.get('cost', ''),
            suggest.get('turnaround', '')
        ]
        ws.append(row)

        # Aplica fonte monospace para condições
        when_cell = ws.cell(row=ws.max_row, column=3)
        when_cell.font = MONOSPACE_FONT

    auto_adjust_columns(ws)
    return len(triggers)

def create_cutoffs_sheet(wb, data):
    """Cria aba Cutoffs"""
    ws = wb.create_sheet("Cutoffs")

    # Headers
    headers = ["Parameter", "Age Group", "Sex", "Low Value", "Critical Low",
               "High Value", "Critical High", "Unit"]
    ws.append(headers)
    apply_header_style(ws)

    cutoffs = data.get('cutoffs', {})

    # Processar cutoffs
    for param, values in cutoffs.items():
        if isinstance(values, dict):
            for age_sex, value in values.items():
                if isinstance(value, dict):
                    row = [
                        param,
                        age_sex,
                        "",
                        value.get('low', ''),
                        value.get('critical_low', ''),
                        value.get('high', ''),
                        value.get('critical_high', ''),
                        value.get('unit', '')
                    ]
                    ws.append(row)
                else:
                    row = [param, age_sex, "", str(value), "", "", "", ""]
                    ws.append(row)

    auto_adjust_columns(ws)

def create_schema_sheet(wb, data):
    """Cria aba Schema CBC"""
    ws = wb.create_sheet("Schema CBC")

    # Headers
    headers = ["Field Name", "Type", "Unit", "Required", "Range", "Description"]
    ws.append(headers)
    apply_header_style(ws)

    fields = data.get('fields', [])

    for field in fields:
        row = [
            field.get('name', ''),
            field.get('type', ''),
            field.get('unit', ''),
            "YES" if field.get('required') else "NO",
            field.get('range', ''),
            field.get('description', '')
        ]
        ws.append(row)

    auto_adjust_columns(ws)

def create_metadata_sheet(wb, yaml_files):
    """Cria aba Metadados"""
    ws = wb.create_sheet("Metadados")

    # Headers
    headers = ["Arquivo YAML", "Versão", "Última Modificação", "Tamanho (linhas)", "Elementos"]
    ws.append(headers)
    apply_header_style(ws)

    for yaml_file in yaml_files:
        filepath = YAML_DIR / yaml_file
        if filepath.exists():
            with open(filepath, 'r', encoding='utf-8') as f:
                lines = len(f.readlines())

            data = load_yaml(yaml_file)
            version = data.get('version', 'N/A')

            # Contar elementos
            elements = 0
            for key, value in data.items():
                if isinstance(value, list):
                    elements += len(value)

            row = [
                yaml_file,
                version,
                datetime.fromtimestamp(filepath.stat().st_mtime).strftime('%Y-%m-%d %H:%M'),
                lines,
                elements
            ]
            ws.append(row)

    auto_adjust_columns(ws)

def main():
    print("🔧 HemoDoctor Excel Generator v2.4.0")
    print("=" * 60)

    # Carrega YAMLs
    print("\n📂 Carregando YAMLs...")
    evidences_data = load_yaml("02_evidence_hybrid.yaml")
    syndromes_data = load_yaml("03_syndromes_hybrid.yaml")
    next_steps_data = load_yaml("09_next_steps_engine_hybrid.yaml")
    config_data = load_yaml("00_config_hybrid.yaml")
    schema_data = load_yaml("01_schema_hybrid.yaml")

    # Cria workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove sheet padrão

    # Estatísticas
    print("\n📊 Processando dados...")

    # Cria sheets
    print("  - Criando aba Evidências...")
    total_evidences = create_evidences_sheet(wb, evidences_data)

    print("  - Criando aba Síndromes...")
    total_syndromes = create_syndromes_sheet(wb, syndromes_data)

    print("  - Criando aba Next Steps...")
    total_triggers = create_next_steps_sheet(wb, next_steps_data)

    print("  - Criando aba Cutoffs...")
    create_cutoffs_sheet(wb, config_data)

    print("  - Criando aba Schema CBC...")
    create_schema_sheet(wb, schema_data)

    # Estatísticas para resumo
    stats = {
        'total_evidences': total_evidences,
        'total_syndromes': total_syndromes,
        'total_triggers': total_triggers,
        'total_fields': len(schema_data.get('fields', [])),
        'critical_evidences': len(evidences_data.get('critical_evidences', [])),
        'strong_evidences': sum(1 for ev in extract_all_evidences(evidences_data) if ev.get('strength') == 'strong'),
        'moderate_evidences': sum(1 for ev in extract_all_evidences(evidences_data) if ev.get('strength') == 'moderate'),
        'weak_evidences': sum(1 for ev in extract_all_evidences(evidences_data) if ev.get('strength') == 'weak'),
        'critical_syndromes': len(syndromes_data.get('critical_syndromes', [])),
        'priority_syndromes': len(syndromes_data.get('priority_syndromes', [])),
        'review_syndromes': len(syndromes_data.get('review_sample_syndromes', [])),
        'routine_syndromes': len(syndromes_data.get('routine_syndromes', [])),
    }

    print("  - Criando aba Resumo...")
    create_summary_sheet(wb, stats)

    print("  - Criando aba Metadados...")
    yaml_files = [
        "00_config_hybrid.yaml",
        "01_schema_hybrid.yaml",
        "02_evidence_hybrid.yaml",
        "03_syndromes_hybrid.yaml",
        "09_next_steps_engine_hybrid.yaml"
    ]
    create_metadata_sheet(wb, yaml_files)

    # Salva
    print(f"\n💾 Salvando arquivo: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)

    # Relatório final
    print("\n✅ EXCEL CRIADO COM SUCESSO!")
    print("=" * 60)
    print(f"\n📄 Arquivo: {OUTPUT_FILE}")
    print(f"📊 Tamanho: {OUTPUT_FILE.stat().st_size / 1024:.1f} KB")
    print(f"\n📋 Conteúdo:")
    print(f"  - {stats['total_evidences']} evidências")
    print(f"  - {stats['total_syndromes']} síndromes")
    print(f"  - {stats['total_triggers']} triggers")
    print(f"  - {stats['total_fields']} campos CBC")
    print(f"\n🎯 Abas criadas:")
    for sheet in wb.sheetnames:
        print(f"  ✓ {sheet}")
    print("\n🎉 Processamento completo!")

if __name__ == "__main__":
    main()
